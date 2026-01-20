#!/usr/bin/env python3

from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
import requests
import argparse
import base64
import os
import re
import json
import hashlib
from datetime import datetime, timedelta, date
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io
from requests import Session
from planning import Issue, ScenarioConfig, compute_slack, schedule_issues

# Load environment variables from .env file
load_dotenv()

app = Flask(__name__)
CORS(app)  # Enable CORS for all routes

# Reuse a single HTTP session to avoid reconnect overhead on repeated calls
HTTP_SESSION = Session()

# CONFIGURATION - Load from environment variables
JIRA_URL = os.getenv('JIRA_URL')
JIRA_EMAIL = os.getenv('JIRA_EMAIL')
JIRA_TOKEN = os.getenv('JIRA_TOKEN')
JQL_QUERY = os.getenv('JQL_QUERY', 'project IN (PRODUCT, TECH) ORDER BY created DESC')
JIRA_BOARD_ID = os.getenv('JIRA_BOARD_ID')  # Optional: board ID for faster sprint fetching
JIRA_TEAM_FIELD_ID = os.getenv('JIRA_TEAM_FIELD_ID', 'customfield_30101')  # Optional: custom field id for Team[Team]
JIRA_TEAM_FALLBACK_FIELD_ID = 'customfield_30101'
JIRA_EPIC_LINK_FIELD_ID = os.getenv('JIRA_EPIC_LINK_FIELD_ID', '').strip()  # Optional: custom field id for Epic Link
JIRA_PRODUCT_PROJECT = os.getenv('JIRA_PRODUCT_PROJECT', 'PRODUCT ROADMAPS')
JIRA_TECH_PROJECT = os.getenv('JIRA_TECH_PROJECT', 'TECHNICAL ROADMAP')
SERVER_PORT = int(os.getenv('SERVER_PORT', '5050'))
EPIC_EMPTY_EXCLUDED_STATUSES = [s.strip() for s in os.getenv('EPIC_EMPTY_EXCLUDED_STATUSES', 'Killed,Done,Incomplete').split(',') if s.strip()]
EPIC_EMPTY_TEAM_IDS = [s.strip() for s in os.getenv('EPIC_EMPTY_TEAM_IDS', '').split(',') if s.strip()]
MISSING_INFO_COMPONENT = os.getenv('MISSING_INFO_COMPONENT', '').strip()
MISSING_INFO_TEAM_IDS = [s.strip() for s in os.getenv('MISSING_INFO_TEAM_IDS', '').split(',') if s.strip()]
STATS_JQL_BASE = os.getenv('STATS_JQL_BASE', '').strip()
STATS_JQL_ORDER_BY = os.getenv('STATS_JQL_ORDER_BY', 'ORDER BY cf[30101] ASC, status ASC').strip()
STATS_PRODUCT_PROJECTS = [s.strip() for s in os.getenv('STATS_PRODUCT_PROJECTS', JIRA_PRODUCT_PROJECT).split(',') if s.strip()]
STATS_TECH_PROJECTS = [s.strip() for s in os.getenv('STATS_TECH_PROJECTS', JIRA_TECH_PROJECT).split(',') if s.strip()]
STATS_EXAMPLE_FILE = os.getenv('STATS_EXAMPLE_FILE', '2025q3.example.json').strip()
STATS_TEAM_IDS = [s.strip() for s in os.getenv('STATS_TEAM_IDS', '').split(',') if s.strip()]
CAPACITY_PROJECT = os.getenv('CAPACITY_PROJECT', '').strip()
CAPACITY_FIELD_ID = os.getenv('CAPACITY_FIELD_ID', '').strip()
CAPACITY_FIELD_NAME = os.getenv('CAPACITY_FIELD_NAME', 'Team capacity[Number]').strip()

SCENARIO_CACHE = {'generatedAt': None, 'data': None}

# Cache settings
SPRINTS_CACHE_FILE = 'sprints_cache.json'
STATS_CACHE_FILE = 'stats_cache.json'
CACHE_EXPIRY_HOURS = 24

def parse_args():
    """Parse CLI arguments to optionally override environment variables."""
    parser = argparse.ArgumentParser(description='Jira proxy server')
    parser.add_argument('--server_port', type=int, help='Port to run the server on (defaults to 5050 or SERVER_PORT env)')
    parser.add_argument('--jira_email', help='Jira account email (overrides JIRA_EMAIL env)')
    parser.add_argument('--jira_token', help='Jira API token (overrides JIRA_TOKEN env)')
    parser.add_argument('--jira_url', help='Base Jira URL, e.g. https://your-domain.atlassian.net (overrides JIRA_URL env)')
    parser.add_argument('--jira_query', help='JQL query to use for fetching issues (overrides JQL_QUERY env)')
    return parser.parse_args()


def add_clause_to_jql(jql: str, clause: str) -> str:
    """Append a clause to JQL before ORDER BY if present."""
    if not clause:
        return jql

    if 'ORDER BY' in jql:
        parts = jql.split('ORDER BY')
        return f"{parts[0].strip()} AND {clause} ORDER BY {parts[1].strip()}"
    return f"{jql} AND {clause}"


def parse_iso_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except Exception:
        return None


def resolve_sprint_label(sprint_value):
    if sprint_value is None:
        return None
    sprint_str = str(sprint_value).strip()
    if not sprint_str:
        return None
    if sprint_str.isdigit():
        cache = load_sprints_cache() or {}
        for sprint in cache.get('sprints', []) or []:
            if str(sprint.get('id')) == sprint_str:
                return sprint.get('name') or sprint_str
    return sprint_str


def quarter_dates_from_label(label):
    if not label:
        return None, None
    match = re.match(r'^(\d{4})Q([1-4])$', str(label).strip(), re.IGNORECASE)
    if not match:
        return None, None
    year = int(match.group(1))
    quarter = int(match.group(2))
    if quarter == 1:
        return date(year, 1, 1), date(year, 3, 31)
    if quarter == 2:
        return date(year, 4, 1), date(year, 6, 30)
    if quarter == 3:
        return date(year, 7, 1), date(year, 9, 30)
    return date(year, 10, 1), date(year, 12, 31)


TEAM_FIELD_CACHE = None
EPIC_LINK_FIELD_CACHE = None
CAPACITY_FIELD_CACHE = None


def resolve_team_field_id(headers):
    """Resolve the Jira custom field ID for Team[Team]."""
    global TEAM_FIELD_CACHE
    if TEAM_FIELD_CACHE:
        return TEAM_FIELD_CACHE
    if JIRA_TEAM_FIELD_ID:
        TEAM_FIELD_CACHE = JIRA_TEAM_FIELD_ID
        return TEAM_FIELD_CACHE

    try:
        response = requests.get(f'{JIRA_URL}/rest/api/3/field', headers=headers, timeout=20)
        if response.status_code != 200:
            return None

        fields = response.json() or []
        for field in fields:
            name = str(field.get('name', '')).strip().lower()
            if name == 'team[team]':
                TEAM_FIELD_CACHE = field.get('id')
                return TEAM_FIELD_CACHE
    except Exception:
        return None

    return None


def resolve_epic_link_field_id(headers, names_map=None):
    """Resolve the Jira custom field ID for Epic Link."""
    global EPIC_LINK_FIELD_CACHE
    if EPIC_LINK_FIELD_CACHE:
        return EPIC_LINK_FIELD_CACHE
    if JIRA_EPIC_LINK_FIELD_ID:
        EPIC_LINK_FIELD_CACHE = JIRA_EPIC_LINK_FIELD_ID
        return EPIC_LINK_FIELD_CACHE

    if names_map:
        for field_id, field_name in (names_map or {}).items():
            if str(field_name).strip().lower() == 'epic link':
                EPIC_LINK_FIELD_CACHE = field_id
                return EPIC_LINK_FIELD_CACHE

    try:
        response = requests.get(f'{JIRA_URL}/rest/api/3/field', headers=headers, timeout=20)
        if response.status_code != 200:
            return None

        fields = response.json() or []
        for field in fields:
            name = str(field.get('name', '')).strip().lower()
            if name == 'epic link':
                EPIC_LINK_FIELD_CACHE = field.get('id')
                return EPIC_LINK_FIELD_CACHE
    except Exception:
        return None

    return None


def resolve_capacity_field_id(headers):
    """Resolve the Jira custom field ID for Team capacity."""
    global CAPACITY_FIELD_CACHE
    if CAPACITY_FIELD_CACHE:
        return CAPACITY_FIELD_CACHE
    if CAPACITY_FIELD_ID:
        CAPACITY_FIELD_CACHE = CAPACITY_FIELD_ID
        return CAPACITY_FIELD_CACHE

    if not CAPACITY_FIELD_NAME:
        return None

    try:
        response = requests.get(f'{JIRA_URL}/rest/api/3/field', headers=headers, timeout=20)
        if response.status_code != 200:
            return None

        fields = response.json() or []
        target = CAPACITY_FIELD_NAME.strip().lower()
        for field in fields:
            name = str(field.get('name', '')).strip().lower()
            if name == target:
                CAPACITY_FIELD_CACHE = field.get('id')
                return CAPACITY_FIELD_CACHE
    except Exception:
        return None

    return None


def extract_team_name(value):
    """Extract a readable team name from Jira Team field values."""
    if value is None:
        return None
    if isinstance(value, list):
        names = [extract_team_name(item) for item in value]
        names = [name for name in names if name]
        return ', '.join(names) if names else None
    if isinstance(value, dict):
        for key in ('name', 'title', 'value', 'displayName', 'teamName'):
            if value.get(key):
                return value.get(key)
        return value.get('id')
    return str(value)


def extract_team_ids(value):
    """Extract Team[Team] ids from Jira Team field values."""
    if value is None:
        return []
    if isinstance(value, list):
        ids = []
        for item in value:
            ids.extend(extract_team_ids(item))
        return [team_id for team_id in ids if team_id]
    if isinstance(value, dict):
        team_id = value.get('id') or value.get('teamId')
        return [team_id] if team_id else []
    return []


def normalize_team_value(value):
    """Normalize Team field values to human-readable names."""
    if isinstance(value, list):
        return [normalize_team_value(item) for item in value if item]
    if isinstance(value, dict):
        return value.get('name') or value.get('value') or value.get('displayName') or value.get('teamName') or value.get('title') or value.get('id')
    return value


def normalize_capacity_team_name(team_name):
    """Strip prefixes to match capacity team labels."""
    if not team_name:
        return None
    cleaned = str(team_name).replace('\u00a0', ' ').strip()
    cleaned = re.sub(r'^\[archived\]\s*', '', cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r'^r&d\s+', '', cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r'^(product|tech)\s*-\s*', '', cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r'\s+', ' ', cleaned)
    return cleaned.strip()


def build_team_value(raw_team):
    """Build a consistent team payload with id/name when possible."""
    if isinstance(raw_team, dict):
        return {
            'id': raw_team.get('id') or raw_team.get('teamId'),
            'name': raw_team.get('name') or raw_team.get('title') or raw_team.get('value') or raw_team.get('displayName')
        }
    return raw_team


def jira_search_request(headers, payload):
    """Call Jira search endpoint using query parameters for /search/jql."""
    url = f'{JIRA_URL}/rest/api/3/search/jql'
    params = {}

    def to_csv(value):
        if isinstance(value, list):
            return ','.join(value)
        return value

    for key in ('jql', 'startAt', 'maxResults', 'expand', 'fields', 'fieldsByKeys'):
        if key in payload and payload[key] is not None:
            params[key] = to_csv(payload[key])

    return HTTP_SESSION.get(url, params=params, headers=headers, timeout=30)


def build_capacity_jql(sprint_name, team_names=None):
    sprint_label = str(sprint_name or '').replace('"', '\\"')
    if team_names:
        clauses = []
        for name in team_names:
            cleaned = str(name).replace('"', '\\"').strip()
            if not cleaned:
                continue
            phrase = f'\\"Team info {sprint_label} - {cleaned}\\"'
            clauses.append(f'summary ~ "{phrase}"')
        if clauses:
            return f'project = "{CAPACITY_PROJECT}" AND ({ " OR ".join(clauses) })'
    phrase = f'\\"Team info {sprint_label} -\\"'
    return f'project = "{CAPACITY_PROJECT}" AND summary ~ "{phrase}"'


def fetch_capacity_for_sprint(sprint_name, headers, debug=False, team_names=None):
    if not CAPACITY_PROJECT:
        return {
            'enabled': False,
            'capacities': {}
        }, None

    capacity_field_id = resolve_capacity_field_id(headers)
    if not capacity_field_id:
        return {
            'enabled': False,
            'capacities': {},
            'message': 'Missing Team capacity field ID'
        }, None

    capacities = {}
    debug_items = []
    issues = []
    jqls = []
    chunk_size = 20
    team_chunks = [team_names[i:i + chunk_size] for i in range(0, len(team_names or []), chunk_size)]
    if not team_chunks:
        team_chunks = [None]

    for chunk in team_chunks:
        jql = build_capacity_jql(sprint_name, chunk)
        jqls.append(jql)
        payload = {
            'jql': jql,
            'startAt': 0,
            'maxResults': 200,
            'fields': ['summary', capacity_field_id]
        }
        response = jira_search_request(headers, payload)
        if response.status_code != 200:
            return None, response.text
        data = response.json() or {}
        issues.extend(data.get('issues') or [])
    pattern = re.compile(rf'^Team info\s+{re.escape(str(sprint_name))}\s*-\s*(.+)$', re.IGNORECASE)
    for issue in issues:
        fields = issue.get('fields') or {}
        summary = str(fields.get('summary') or '').strip()
        match = pattern.match(summary)
        if not match:
            continue
        short_name = normalize_capacity_team_name(match.group(1))
        if not short_name:
            continue
        raw_capacity = fields.get(capacity_field_id)
        if debug:
            debug_items.append({
                'summary': summary,
                'rawCapacity': raw_capacity
            })
        try:
            capacity_value = float(raw_capacity)
        except (TypeError, ValueError):
            continue
        capacities[short_name] = capacity_value

    response_payload = {
        'enabled': True,
        'sprint': sprint_name,
        'capacities': capacities
    }
    if debug:
        response_payload['debug'] = {
            'jql': jqls if len(jqls) > 1 else jqls[0],
            'issueCount': len(issues),
            'matched': debug_items[:20],
            'fieldId': capacity_field_id
        }
    return response_payload, None


def fetch_watchers_count(issue_key, headers):
    """Fetch watchers count for an issue (fallback if watches field is missing)."""
    if not issue_key:
        return None
    try:
        response = requests.get(
            f'{JIRA_URL}/rest/api/3/issue/{issue_key}/watchers',
            headers=headers,
            timeout=20
        )
        if response.status_code != 200:
            print(f'❌ Watchers fetch error for {issue_key}: {response.status_code} {response.text}')
            return None
        data = response.json() or {}
        if isinstance(data.get('watchCount'), int):
            return data['watchCount']
        watchers = data.get('watchers') or []
        return len(watchers)
    except Exception as e:
        print(f'❌ Watchers fetch exception for {issue_key}: {e}')
        return None


def fetch_capacity_team_sizes(sprint_name, headers, team_names=None):
    """Fetch team sizes from Jira capacity issues (watchers count)."""
    if not CAPACITY_PROJECT or not sprint_name:
        return {}, {}

    issues = []
    jqls = []
    chunk_size = 20
    team_chunks = [team_names[i:i + chunk_size] for i in range(0, len(team_names or []), chunk_size)]
    if not team_chunks:
        team_chunks = [None]

    for chunk in team_chunks:
        jql = build_capacity_jql(sprint_name, chunk)
        jqls.append(jql)
        payload = {
            'jql': jql,
            'startAt': 0,
            'maxResults': 200,
            'fields': ['summary', 'watches']
        }
        response = jira_search_request(headers, payload)
        if response.status_code != 200:
            print(f'❌ Capacity size fetch error: {response.status_code} {response.text}')
            continue
        data = response.json() or {}
        issues.extend(data.get('issues') or [])

    sizes = {}
    details = {}
    pattern = re.compile(rf'^Team info\s+{re.escape(str(sprint_name))}\s*-\s*(.+)$', re.IGNORECASE)
    for issue in issues:
        fields = issue.get('fields') or {}
        summary = str(fields.get('summary') or '').strip()
        match = pattern.match(summary)
        if not match:
            continue
        short_name = normalize_capacity_team_name(match.group(1))
        if not short_name:
            continue
        watch_count = None
        watches = fields.get('watches') or {}
        if isinstance(watches, dict):
            watch_count = watches.get('watchCount')
        if watch_count is None:
            watch_count = fetch_watchers_count(issue.get('key'), headers)
        if watch_count is None:
            continue
        try:
            count = int(watch_count)
            sizes[short_name] = count
            details[short_name] = {
                'watchers': count,
                'issue_key': issue.get('key')
            }
            if issue.get('key'):
                print(f'🧭 Capacity size: {short_name} -> {issue.get("key")} watchers={count}')
        except (TypeError, ValueError):
            continue

    return sizes, details


# Cache helper functions
def load_sprints_cache():
    """Load sprints from cache file"""
    try:
        if os.path.exists(SPRINTS_CACHE_FILE):
            with open(SPRINTS_CACHE_FILE, 'r') as f:
                cache_data = json.load(f)
                return cache_data
        return None
    except Exception as e:
        print(f'⚠️ Failed to load cache: {e}')
        return None


def save_sprints_cache(sprints):
    """Save sprints to cache file"""
    try:
        cache_data = {
            'timestamp': datetime.now().isoformat(),
            'sprints': sprints
        }
        with open(SPRINTS_CACHE_FILE, 'w') as f:
            json.dump(cache_data, f, indent=2)
        print(f'💾 Cached {len(sprints)} sprints to {SPRINTS_CACHE_FILE}')
        return True
    except Exception as e:
        print(f'⚠️ Failed to save cache: {e}')
        return False


def is_cache_valid():
    """Check if cache exists and is not expired"""
    cache_data = load_sprints_cache()
    if not cache_data or 'timestamp' not in cache_data:
        return False

    try:
        cache_time = datetime.fromisoformat(cache_data['timestamp'])
        expiry_time = cache_time + timedelta(hours=CACHE_EXPIRY_HOURS)
        is_valid = datetime.now() < expiry_time

        if is_valid:
            hours_old = (datetime.now() - cache_time).total_seconds() / 3600
            print(f'✅ Cache is valid (age: {hours_old:.1f} hours)')
        else:
            print(f'⏰ Cache expired (age: {(datetime.now() - cache_time).total_seconds() / 3600:.1f} hours)')

        return is_valid
    except Exception as e:
        print(f'⚠️ Failed to validate cache: {e}')
        return False


def load_stats_cache():
    """Load stats cache from disk."""
    try:
        if os.path.exists(STATS_CACHE_FILE):
            with open(STATS_CACHE_FILE, 'r') as f:
                return json.load(f)
        return {}
    except Exception as e:
        print(f'⚠️ Failed to load stats cache: {e}')
        return {}


def save_stats_cache(cache_data):
    """Persist stats cache to disk."""
    try:
        with open(STATS_CACHE_FILE, 'w') as f:
            json.dump(cache_data, f, indent=2)
        return True
    except Exception as e:
        print(f'⚠️ Failed to save stats cache: {e}')
        return False


def build_stats_cache_key(sprint_name, base_jql, team_ids):
    raw = f"{sprint_name}::{base_jql}::{','.join(team_ids or [])}::{STATS_JQL_ORDER_BY}"
    digest = hashlib.sha1(raw.encode('utf-8')).hexdigest()[:12]
    return f"sprint:{sprint_name}:{digest}"


def strip_sprint_clause(jql):
    """Remove Sprint clause if the base query already includes it."""
    if not jql:
        return jql
    jql = re.sub(r'\s+AND\s+Sprint\s+in\s+\([^)]+\)', '', jql, flags=re.IGNORECASE)
    return re.sub(r'\s+AND\s+Sprint\s*=\s*[^ ]+', '', jql, flags=re.IGNORECASE)


def extract_team_ids_from_jql(jql):
    """Extract Team[Team] ids from a JQL clause if present."""
    if not jql:
        return []
    match_in = re.search(r'"Team\[Team\]"\s+in\s*\(([^)]+)\)', jql, flags=re.IGNORECASE)
    if match_in:
        raw = match_in.group(1)
        parts = [p.strip() for p in raw.split(',')]
        ids = []
        for part in parts:
            part = part.strip().strip('"').strip("'")
            if part:
                ids.append(part)
        return ids
    match_eq = re.search(r'"Team\[Team\]"\s*=\s*("?)([^")\s]+)\1', jql, flags=re.IGNORECASE)
    if match_eq:
        return [match_eq.group(2)]
    return []


def get_stats_team_ids():
    """Resolve stats team IDs from env or JQL configuration."""
    if STATS_TEAM_IDS:
        return STATS_TEAM_IDS
    base_jql = STATS_JQL_BASE or JQL_QUERY
    return extract_team_ids_from_jql(base_jql)


def classify_project(project_name):
    """Classify projects into product/tech buckets based on config."""
    if not project_name:
        return 'other'
    normalized = str(project_name).strip().lower()
    if any(normalized == p.lower() for p in STATS_PRODUCT_PROJECTS):
        return 'product'
    if any(normalized == p.lower() for p in STATS_TECH_PROJECTS):
        return 'tech'
    return 'other'


def fetch_sprints_from_jira():
    """Fetch sprints from Jira (not from cache)"""
    auth_string = f"{JIRA_EMAIL}:{JIRA_TOKEN}"
    auth_bytes = auth_string.encode('ascii')
    auth_base64 = base64.b64encode(auth_bytes).decode('ascii')

    headers = {
        'Authorization': f'Basic {auth_base64}',
        'Accept': 'application/json',
        'Content-Type': 'application/json'
    }

    formatted_sprints = []

    # Method 1: Try to get sprints from board (if JIRA_BOARD_ID is set)
    if JIRA_BOARD_ID:
        try:
            print(f'\n📅 Fetching sprints from board {JIRA_BOARD_ID}...')
            start_at = 0
            while True:
                response = requests.get(
                    f'{JIRA_URL}/rest/agile/1.0/board/{JIRA_BOARD_ID}/sprint',
                    headers=headers,
                    params={'maxResults': 100, 'startAt': start_at, 'state': 'active,future,closed'},
                    timeout=30
                )

                if response.status_code != 200:
                    print(f'⚠️ Board API returned {response.status_code}, trying alternative method...')
                    break

                data = response.json()
                sprints = data.get('values', [])

                for sprint in sprints:
                    name = sprint.get('name', '')
                    sprint_id = sprint.get('id')
                    state = sprint.get('state', '')

                    if re.match(r'^\d{4}Q[1-4]$', name):
                        formatted_sprints.append({
                            'id': sprint_id,
                            'name': name,
                            'state': state
                        })

                if data.get('isLast', False) or not sprints:
                    break

                start_at += len(sprints)

            if formatted_sprints:
                print(f'✅ Found {len(formatted_sprints)} sprints from board')
            else:
                print(f'⚠️ Board API returned {response.status_code}, trying alternative method...')
        except Exception as board_error:
            print(f'⚠️ Board API failed: {board_error}, trying alternative method...')

    # Method 2: If board method failed or found no sprints, get sprints from issues
    if len(formatted_sprints) == 0:
        print(f'\n📅 Fetching sprints from issues (alternative method)...')

        # Build JQL query without sprint filter to get all issues
        base_jql = STATS_JQL_BASE or f'project in ("{JIRA_PRODUCT_PROJECT}","{JIRA_TECH_PROJECT}")'
        # Remove any existing sprint filters from the query
        base_jql = strip_sprint_clause(base_jql)

        def collect_sprints_by_jql(jql_query, sprints_dict):
            total_issues = 0
            start_at = 0
            while True:
                payload = {
                    'jql': jql_query,
                    'startAt': start_at,
                    'maxResults': 200,  # Reduced from 1000 for better performance
                    'fields': ['customfield_10101']  # Only get sprint field
                }

                response = jira_search_request(headers, payload)
                if response.status_code != 200:
                    break

                data = response.json()
                issues = data.get('issues', [])

                for issue in issues:
                    sprint_field = issue.get('fields', {}).get('customfield_10101', [])
                    if sprint_field and isinstance(sprint_field, list):
                        for sprint in sprint_field:
                            if sprint and isinstance(sprint, dict):
                                name = sprint.get('name', '')
                                sprint_id = sprint.get('id')
                                state = sprint.get('state', '')

                                if re.match(r'^\d{4}Q[1-4]$', name) and sprint_id:
                                    sprints_dict[sprint_id] = {
                                        'id': sprint_id,
                                        'name': name,
                                        'state': state
                                    }

                total_issues += len(issues)
                if len(issues) < payload['maxResults']:
                    break

                start_at += len(issues)

            return total_issues

        sprints_dict = {}
        issues_count = collect_sprints_by_jql(base_jql, sprints_dict)
        closed_jql = add_clause_to_jql(base_jql, 'Sprint in closedSprints()')
        issues_count += collect_sprints_by_jql(closed_jql, sprints_dict)
        future_jql = add_clause_to_jql(base_jql, 'Sprint in futureSprints()')
        issues_count += collect_sprints_by_jql(future_jql, sprints_dict)
        open_jql = add_clause_to_jql(base_jql, 'Sprint in openSprints()')
        issues_count += collect_sprints_by_jql(open_jql, sprints_dict)

        formatted_sprints = list(sprints_dict.values())
        print(f'✅ Found {len(formatted_sprints)} unique sprints from {issues_count} issues')

    # Sort sprints by name (latest first)
    formatted_sprints.sort(key=lambda x: x['name'], reverse=True)

    return formatted_sprints


def fetch_epic_details_bulk(epic_keys, headers, epic_name_field):
    """Fetch epic details in small batches to avoid per-epic network calls."""
    epic_details = {}
    if not epic_keys:
        return epic_details

    epic_field = epic_name_field or 'customfield_10011'
    keys_list = list(epic_keys)
    batch_size = 40  # keep JQL length reasonable for GET

    for start in range(0, len(keys_list), batch_size):
        batch_keys = keys_list[start:start + batch_size]
        jql = f'issueKey in ({",".join(batch_keys)})'
        payload = {
            'jql': jql,
            'maxResults': len(batch_keys),
            'fields': ['summary', 'reporter', 'assignee', epic_field]
        }

        try:
            resp = jira_search_request(headers, payload)
            if resp.status_code != 200:
                print(f'⚠️ Epic batch {start}-{start + len(batch_keys)} failed: {resp.status_code}')
                continue

            data = resp.json()
            for issue in data.get('issues', []):
                fields = issue.get('fields', {}) or {}
                key = issue.get('key')
                epic_details[key] = {
                    'key': key,
                    'summary': fields.get('summary'),
                    'reporter': (fields.get('reporter') or {}).get('displayName'),
                    'assignee': {'displayName': (fields.get('assignee') or {}).get('displayName')} if fields.get('assignee') else None,
                }
        except Exception as exc:
            print(f'⚠️ Epic batch fetch error: {exc}')

    return epic_details


def derive_epic_jql(base_jql: str, team_ids=None) -> str:
    """Attempt to derive an epic query from a story query by swapping Story→Epic."""
    if not base_jql:
        base_jql = ''

    jql = base_jql
    replacements = [
        ('type = Story', 'type = Epic'),
        ('type=Story', 'type=Epic'),
        ('type = "Story"', 'type = "Epic"'),
        ("type = 'Story'", "type = 'Epic'"),
        ('issuetype = Story', 'issuetype = Epic'),
        ('issuetype=Story', 'issuetype=Epic'),
        ('issuetype = "Story"', 'issuetype = "Epic"'),
        ("issuetype = 'Story'", "issuetype = 'Epic'"),
    ]
    replaced = False
    for old, new in replacements:
        if old in jql:
            jql = jql.replace(old, new)
            replaced = True

    if not replaced:
        jql = add_clause_to_jql(jql, 'type = Epic') if jql else 'type = Epic'

    if EPIC_EMPTY_EXCLUDED_STATUSES:
        quoted = ', '.join(f'"{s}"' for s in EPIC_EMPTY_EXCLUDED_STATUSES)
        jql = add_clause_to_jql(jql, f'status not in ({quoted})')

    team_ids = [t.strip() for t in (team_ids or []) if t and str(t).strip()]
    if team_ids:
        if len(team_ids) == 1:
            jql = add_clause_to_jql(jql, f'"Team[Team]" = "{team_ids[0]}"')
        else:
            quoted_teams = ', '.join(f'"{t}"' for t in team_ids)
            jql = add_clause_to_jql(jql, f'"Team[Team]" in ({quoted_teams})')
    return jql


def fetch_epics_for_empty_alert(jql, headers, team_field_id, epic_name_field):
    """Fetch epics matching the current sprint/team filters so UI can flag epics with 0 stories."""
    epic_jql = derive_epic_jql(jql, EPIC_EMPTY_TEAM_IDS)
    epic_field = epic_name_field or 'customfield_10011'

    fields_list = ['summary', 'status', 'assignee', epic_field]
    if team_field_id and team_field_id not in fields_list:
        fields_list.append(team_field_id)
    if JIRA_TEAM_FALLBACK_FIELD_ID not in fields_list:
        fields_list.append(JIRA_TEAM_FALLBACK_FIELD_ID)

    payload = {
        'jql': epic_jql,
        'startAt': 0,
        'maxResults': 250,
        'fields': fields_list
    }
    resp = jira_search_request(headers, payload)
    if resp.status_code != 200:
        print(f'⚠️ Epic empty-state fetch failed: {resp.status_code}')
        return []

    data = resp.json() or {}
    issues = data.get('issues', []) or []
    epics = []
    for issue in issues:
        fields = issue.get('fields', {}) or {}

        raw_team = None
        if fields.get(JIRA_TEAM_FALLBACK_FIELD_ID) is not None:
            raw_team = fields.get(JIRA_TEAM_FALLBACK_FIELD_ID)
        elif team_field_id and fields.get(team_field_id) is not None:
            raw_team = fields.get(team_field_id)

        team_value = build_team_value(raw_team) if raw_team is not None else None
        team_name = extract_team_name(raw_team) if raw_team is not None else None

        assignee = fields.get('assignee') or {}
        status = fields.get('status') or {}
        epics.append({
            'key': issue.get('key'),
            'summary': fields.get('summary'),
            'status': {'name': status.get('name')} if status else None,
            'assignee': {'displayName': assignee.get('displayName')} if assignee else None,
            'team': team_value,
            'teamName': team_name,
            'teamId': team_value.get('id') if isinstance(team_value, dict) else None,
        })
    return epics


def fetch_story_counts_for_epics(epic_keys, headers, epic_link_field):
    """Return total Story counts for each epic key.

    Prefer counting via the Epic Link field (company-managed Jira). If that yields zero for an epic,
    fall back to counting via `parent` (team-managed projects / some Jira configs).
    """
    epic_keys = [k for k in (epic_keys or []) if k]
    if not epic_keys:
        return {}

    def count_by_query(batch_keys, jql, parse_epic_key, fields):
        start_at = 0
        max_results = 250
        local_counts = {k: 0 for k in batch_keys}

        while True:
            payload = {
                'jql': jql,
                'startAt': start_at,
                'maxResults': max_results,
                'fields': fields
            }
            resp = jira_search_request(headers, payload)
            if resp.status_code != 200:
                return local_counts

            data = resp.json() or {}
            issues = data.get('issues', []) or []
            if not issues:
                break

            for issue in issues:
                fields_obj = issue.get('fields', {}) or {}
                epic_key = parse_epic_key(fields_obj)
                if epic_key in local_counts:
                    local_counts[epic_key] += 1

            start_at += len(issues)
            total = data.get('total')
            if total is not None and start_at >= total:
                break
            if len(issues) < max_results:
                break

        return local_counts

    counts = {k: 0 for k in epic_keys}
    batch_size = 40

    for start in range(0, len(epic_keys), batch_size):
        batch = epic_keys[start:start + batch_size]

        # 1) Try Epic Link counting (company-managed)
        if epic_link_field:
            epic_link_jql = f'"Epic Link" in ({",".join(batch)}) AND issuetype = Story'
            link_counts = count_by_query(
                batch,
                epic_link_jql,
                lambda f: f.get(epic_link_field),
                [epic_link_field]
            )
            for k, v in link_counts.items():
                counts[k] += v

        # 2) Fall back to parent counting for epics that still have 0
        remaining = [k for k in batch if counts.get(k, 0) == 0]
        if remaining:
            parent_jql = f'parent in ({",".join(remaining)}) AND issuetype = Story'
            parent_counts = count_by_query(
                remaining,
                parent_jql,
                lambda f: (f.get('parent') or {}).get('key'),
                ['parent']
            )
            for k, v in parent_counts.items():
                counts[k] += v

    return counts


def fetch_tasks(include_team_name=False):
    """Fetch tasks from Jira API."""
    try:
        # Get sprint parameter from query string
        sprint = request.args.get('sprint', '')
        team = request.args.get('team', '').strip()
        project_filter = request.args.get('project', '').strip().lower()

        # Prepare authorization
        auth_string = f"{JIRA_EMAIL}:{JIRA_TOKEN}"
        auth_bytes = auth_string.encode('ascii')
        auth_base64 = base64.b64encode(auth_bytes).decode('ascii')

        # Prepare headers
        headers = {
            'Authorization': f'Basic {auth_base64}',
            'Accept': 'application/json',
            'Content-Type': 'application/json'
        }

        # Build JQL query with sprint filter if provided
        jql = JQL_QUERY
        if sprint:
            jql = add_clause_to_jql(jql, f"Sprint = {sprint}")

        if team and team.lower() != 'all':
            jql = add_clause_to_jql(jql, f'"Team[Team]" = {team}')

        if project_filter in ('product', 'tech'):
            project_name = JIRA_PRODUCT_PROJECT if project_filter == 'product' else JIRA_TECH_PROJECT
            jql = add_clause_to_jql(jql, f'project = "{project_name}"')

        team_field_id = resolve_team_field_id(headers)
        epic_link_field_id = resolve_epic_link_field_id(headers)

        # Prepare request parameters for search endpoint
        fields_list = [
            'summary',
            'status',
            'priority',
            'issuetype',
            'assignee',
            'updated',
            'customfield_10004',  # Story Points
            'parent'
        ]
        if epic_link_field_id and epic_link_field_id not in fields_list:
            fields_list.append(epic_link_field_id)
        if team_field_id:
            fields_list.append(team_field_id)
        if JIRA_TEAM_FALLBACK_FIELD_ID not in fields_list:
            fields_list.append(JIRA_TEAM_FALLBACK_FIELD_ID)
        if not team_field_id:
            print('⚠️ Team field id not resolved; using customfield_30101 fallback.')

        max_results = 250
        start_at = 0
        collected_issues = []
        names_map = {}
        total_issues = None

        print(f'\n🔍 Making request to Jira API...')
        print(f'URL: {JIRA_URL}/rest/api/3/search/jql')
        print(f'Sprint: {sprint if sprint else "All"}')
        print(f'JQL: {jql}')

        while len(collected_issues) < max_results:
            payload = {
                'jql': jql,
                'startAt': start_at,
                'maxResults': max_results,
                'fields': fields_list
            }

            response = jira_search_request(headers, payload)
            print(f'📊 Response Status: {response.status_code}')

            if response.status_code != 200:
                error_text = response.text
                print(f'❌ Error Response: {error_text}')

                try:
                    error_json = response.json()
                    print(f'Error Details: {error_json}')
                except:
                    pass

                error_response = jsonify({
                    'error': f'Jira API error: {response.status_code}',
                    'details': error_text,
                    'jql_used': jql
                })
                error_response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
                error_response.headers['Pragma'] = 'no-cache'
                error_response.headers['Expires'] = '0'
                return error_response, response.status_code

            data = response.json()
            if not names_map:
                names_map = data.get('names', {}) or {}
            total_issues = data.get('total', total_issues)

            issues = data.get('issues', [])
            if not issues:
                break

            collected_issues.extend(issues)
            # Stop when Jira signals we're at the end or when the last page is smaller than the request size
            if len(issues) < payload['maxResults']:
                break
            if len(collected_issues) >= max_results:
                collected_issues = collected_issues[:max_results]
                break

            start_at += len(issues)
            if total_issues is not None and start_at >= total_issues:
                break

        data = {
            'issues': collected_issues,
            'names': names_map,
            'total': total_issues,
            'startAt': 0,
            'maxResults': max_results
        }

        if total_issues is None:
            total_issues = len(collected_issues)

        if not team_field_id:
            team_field_id = next((k for k, v in names_map.items() if str(v).lower() == 'team[team]'), None)
        epic_link_field = epic_link_field_id or resolve_epic_link_field_id(headers, names_map)
        epic_name_field = next((k for k, v in names_map.items() if str(v).lower() == 'epic name'), None)

        epic_keys = set()

        for issue in collected_issues:
            fields = issue.get('fields', {})

            raw_team = None
            if fields.get(JIRA_TEAM_FALLBACK_FIELD_ID) is not None:
                raw_team = fields.get(JIRA_TEAM_FALLBACK_FIELD_ID)
            elif team_field_id and fields.get(team_field_id) is not None:
                raw_team = fields.get(team_field_id)

            if raw_team is not None:
                team_name = extract_team_name(raw_team)
                fields['team'] = build_team_value(raw_team)
                fields['teamName'] = team_name
                fields['teamId'] = fields['team'].get('id') if isinstance(fields['team'], dict) else None

            parent_fields = fields.get('parent', {}).get('fields', {})
            if parent_fields.get('summary'):
                fields['parentSummary'] = parent_fields.get('summary')

            epic_key = None
            if epic_link_field and fields.get(epic_link_field):
                epic_key = fields.get(epic_link_field)
            elif fields.get('parent') and fields['parent'].get('key') and \
                    fields['parent'].get('fields', {}).get('issuetype', {}).get('name', '').lower() == 'epic':
                epic_key = fields['parent'].get('key')

            if epic_key:
                fields['epicKey'] = epic_key
                epic_keys.add(epic_key)

        epic_details = fetch_epic_details_bulk(epic_keys, headers, epic_name_field)
        epics_in_scope = fetch_epics_for_empty_alert(jql, headers, team_field_id, epic_name_field)
        epic_story_counts = fetch_story_counts_for_epics([e.get('key') for e in epics_in_scope], headers, epic_link_field) if epic_link_field else None
        for epic in epics_in_scope:
            key = epic.get('key')
            epic['totalStories'] = epic_story_counts.get(key) if (epic_story_counts and key) else None
        slim_issues = []
        for issue in collected_issues:
            fields = issue.get('fields', {})
            status = fields.get('status') or {}
            priority = fields.get('priority') or {}
            issuetype = fields.get('issuetype') or {}
            assignee = fields.get('assignee') or {}
            slim_issues.append({
                'id': issue.get('id'),
                'key': issue.get('key'),
                'fields': {
                    'summary': fields.get('summary'),
                    'status': {'name': status.get('name')} if status else None,
                    'priority': {'name': priority.get('name')} if priority else None,
                    'issuetype': {'name': issuetype.get('name')} if issuetype else None,
                    'assignee': {'displayName': assignee.get('displayName')} if assignee else None,
                    'updated': fields.get('updated'),
                    'customfield_10004': fields.get('customfield_10004'),
                    'team': fields.get('team'),
                    'teamName': fields.get('teamName'),
                    'teamId': fields.get('teamId'),
                    'epicKey': fields.get('epicKey'),
                    'parentSummary': fields.get('parentSummary')
                }
            })

        data['issues'] = slim_issues
        data['epics'] = epic_details
        data['epicsInScope'] = epics_in_scope
        data['teamFieldId'] = team_field_id

        print(f'✅ Success! Found {len(data.get("issues", []))} issues')

        success_response = jsonify(data)
        success_response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        success_response.headers['Pragma'] = 'no-cache'
        success_response.headers['Expires'] = '0'
        return success_response
        
    except Exception as e:
        print(f'❌ Exception: {str(e)}')
        import traceback
        traceback.print_exc()
        error_response = jsonify({
            'error': 'Failed to fetch tasks from Jira',
            'message': str(e)
        })
        error_response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        return error_response, 500


def fetch_issues_by_keys(keys, headers, fields_list):
    """Fetch issues by keys in batches."""
    if not keys:
        return []

    results = []
    batch_size = 100
    for i in range(0, len(keys), batch_size):
        batch = keys[i:i + batch_size]
        jql = f'key in ({",".join(batch)})'
        payload = {
            'jql': jql,
            'startAt': 0,
            'maxResults': batch_size,
            'fields': fields_list
        }
        response = jira_search_request(headers, payload)
        if response.status_code != 200:
            print(f'❌ Dependencies fetch error: {response.status_code} {response.text}')
            continue
        data = response.json() or {}
        results.extend(data.get('issues', []) or [])
    return results


def fetch_issues_by_jql(jql, headers, fields_list, max_results=500):
    """Fetch issues by JQL with pagination."""
    results = []
    start_at = 0
    while len(results) < max_results:
        payload = {
            'jql': jql,
            'startAt': start_at,
            'maxResults': min(250, max_results - len(results)),
            'fields': fields_list
        }
        response = jira_search_request(headers, payload)
        if response.status_code != 200:
            print(f'❌ Scenario fetch error: {response.status_code} {response.text}')
            break
        data = response.json() or {}
        issues = data.get('issues', []) or []
        if not issues:
            break
        results.extend(issues)
        start_at += len(issues)
        total = data.get('total')
        if total is not None and start_at >= total:
            break
        if len(issues) < payload['maxResults']:
            break
    return results


def build_scenario_jql(filters):
    jql = JQL_QUERY
    sprint = filters.get('sprint')
    if sprint:
        if str(sprint).isdigit():
            jql = add_clause_to_jql(jql, f"Sprint = {sprint}")
        else:
            jql = add_clause_to_jql(jql, f'Sprint = "{sprint}"')

    teams = [t for t in (filters.get('teams') or []) if t]
    if teams:
        quoted = ', '.join(f'"{t}"' for t in teams)
        jql = add_clause_to_jql(jql, f'"Team[Team]" in ({quoted})')

    projects = [p for p in (filters.get('projects') or []) if p]
    if projects:
        quoted = ', '.join(f'"{p}"' for p in projects)
        jql = add_clause_to_jql(jql, f'project in ({quoted})')

    epics = [e for e in (filters.get('epics') or []) if e]
    if epics:
        quoted = ', '.join(f'"{e}"' for e in epics)
        jql = add_clause_to_jql(jql, f'("Epic Link" in ({quoted}) OR parent in ({quoted}))')

    return jql


def build_issue_snapshot(issue, team_field_id=None, epic_link_field_id=None):
    """Build a compact issue snapshot for dependency rendering."""
    fields = issue.get('fields', {}) or {}
    raw_team = None
    if fields.get(JIRA_TEAM_FALLBACK_FIELD_ID) is not None:
        raw_team = fields.get(JIRA_TEAM_FALLBACK_FIELD_ID)
    elif team_field_id and fields.get(team_field_id) is not None:
        raw_team = fields.get(team_field_id)

    team_name = None
    team_id = None
    if raw_team is not None:
        team_name = extract_team_name(raw_team)
        team_value = build_team_value(raw_team)
        if isinstance(team_value, dict):
            team_id = team_value.get('id')

    epic_key = None
    if epic_link_field_id and fields.get(epic_link_field_id):
        epic_key = fields.get(epic_link_field_id)
    elif fields.get('parent') and fields['parent'].get('key') and \
            fields['parent'].get('fields', {}).get('issuetype', {}).get('name', '').lower() == 'epic':
        epic_key = fields['parent'].get('key')

    return {
        'key': issue.get('key'),
        'summary': fields.get('summary'),
        'issuetype': fields.get('issuetype', {}).get('name') if fields.get('issuetype') else None,
        'status': fields.get('status', {}).get('name') if fields.get('status') else None,
        'priority': fields.get('priority', {}).get('name') if fields.get('priority') else None,
        'storyPoints': fields.get('customfield_10004'),
        'teamName': team_name,
        'teamId': team_id,
        'epicKey': epic_key
    }


def collect_dependencies(keys, headers):
    """Fetch dependency links for a set of issues."""
    keys = sorted({str(k).strip() for k in keys if str(k).strip()})
    if not keys:
        return {}

    def normalize_link_text(value):
        return str(value or '').strip().lower()

    def has_block_marker(*values):
        return any('block' in normalize_link_text(value) for value in values if value)

    def has_depend_marker(*values):
        return any('depend' in normalize_link_text(value) for value in values if value)

    def resolve_link_direction(category, relation, direction, base_key, linked_key):
        relation_text = normalize_link_text(relation)
        if category == 'block':
            base_blocks = 'block' in relation_text and 'blocked by' not in relation_text
            base_blocked = 'blocked by' in relation_text
            if base_blocks:
                return base_key, linked_key
            if base_blocked:
                return linked_key, base_key
            return (base_key, linked_key) if direction == 'outward' else (linked_key, base_key)
        if category == 'dependency':
            base_depends = 'depend' in relation_text and 'depended on' not in relation_text
            base_depended = 'depended on' in relation_text
            if base_depends:
                return linked_key, base_key
            if base_depended:
                return base_key, linked_key
            return (linked_key, base_key) if direction == 'outward' else (base_key, linked_key)
        return None, None

    team_field_id = resolve_team_field_id(headers)
    epic_link_field_id = resolve_epic_link_field_id(headers)

    fields_list = [
        'summary',
        'status',
        'priority',
        'issuetype',
        'customfield_10004',
        'parent',
        'issuelinks'
    ]
    if epic_link_field_id and epic_link_field_id not in fields_list:
        fields_list.append(epic_link_field_id)
    if team_field_id and team_field_id not in fields_list:
        fields_list.append(team_field_id)
    if JIRA_TEAM_FALLBACK_FIELD_ID not in fields_list:
        fields_list.append(JIRA_TEAM_FALLBACK_FIELD_ID)

    issues = fetch_issues_by_keys(keys, headers, fields_list)
    issue_map = {}
    linked_keys = set()
    for issue in issues:
        snapshot = build_issue_snapshot(issue, team_field_id, epic_link_field_id)
        if snapshot.get('key'):
            issue_map[snapshot['key']] = snapshot
        for link in issue.get('fields', {}).get('issuelinks', []) or []:
            linked = link.get('outwardIssue') or link.get('inwardIssue')
            if linked and linked.get('key'):
                linked_keys.add(linked['key'])

    missing_linked = sorted(linked_keys - set(issue_map.keys()))
    if missing_linked:
        linked_issues = fetch_issues_by_keys(missing_linked, headers, fields_list)
        for issue in linked_issues:
            snapshot = build_issue_snapshot(issue, team_field_id, epic_link_field_id)
            if snapshot.get('key'):
                issue_map[snapshot['key']] = snapshot

    dependencies = {}
    for issue in issues:
        base_key = issue.get('key')
        if not base_key:
            continue
        links = issue.get('fields', {}).get('issuelinks', []) or []
        entries = []
        for link in links:
            type_info = link.get('type', {}) or {}
            type_name = type_info.get('name')
            type_inward = type_info.get('inward')
            type_outward = type_info.get('outward')
            linked_issue = link.get('outwardIssue')
            direction = 'outward'
            relation = type_outward
            if linked_issue is None:
                linked_issue = link.get('inwardIssue')
                direction = 'inward'
                relation = type_inward
            if not linked_issue or not linked_issue.get('key'):
                continue
            category = None
            if has_block_marker(type_name, type_inward, type_outward):
                category = 'block'
            elif has_depend_marker(type_name, type_inward, type_outward):
                category = 'dependency'
            if category is None:
                continue
            linked_key = linked_issue.get('key')
            prereq_key, dependent_key = resolve_link_direction(
                category,
                relation,
                direction,
                base_key,
                linked_key
            )
            linked_snapshot = issue_map.get(linked_key)
            if not linked_snapshot:
                linked_snapshot = {
                    'key': linked_key,
                    'summary': linked_issue.get('fields', {}).get('summary'),
                    'issuetype': linked_issue.get('fields', {}).get('issuetype', {}).get('name'),
                    'status': linked_issue.get('fields', {}).get('status', {}).get('name'),
                    'storyPoints': linked_issue.get('fields', {}).get('customfield_10004'),
                    'teamName': None,
                    'teamId': None,
                    'epicKey': None
                }
            entries.append({
                **linked_snapshot,
                'direction': direction,
                'relation': relation or type_name,
                'category': category,
                'typeName': type_name,
                'typeInward': type_inward,
                'typeOutward': type_outward,
                'prereqKey': prereq_key,
                'dependentKey': dependent_key
            })
        if entries:
            dependencies[base_key] = entries

    return dependencies


@app.route('/api/dependencies', methods=['POST'])
def get_dependencies():
    """Fetch dependency links for a set of issues."""
    try:
        payload = request.get_json(silent=True) or {}
        keys = payload.get('keys') or []
        if not keys:
            return jsonify({'dependencies': {}})

        auth_string = f"{JIRA_EMAIL}:{JIRA_TOKEN}"
        auth_bytes = auth_string.encode('ascii')
        auth_base64 = base64.b64encode(auth_bytes).decode('ascii')

        headers = {
            'Authorization': f'Basic {auth_base64}',
            'Accept': 'application/json',
            'Content-Type': 'application/json'
        }

        dependencies = collect_dependencies(keys, headers)
        return jsonify({'dependencies': dependencies})
    except Exception as e:
        print(f'❌ Dependencies error: {e}')
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Failed to fetch dependencies', 'message': str(e)}), 500


@app.route('/api/issues/lookup', methods=['GET'])
def lookup_issues():
    """Lookup issues by key/id for dependency popovers."""
    try:
        keys_param = request.args.get('keys', '') or ''
        ids_param = request.args.get('ids', '') or ''
        keys = [k.strip() for k in keys_param.split(',') if k.strip()]
        ids = [i.strip() for i in ids_param.split(',') if i.strip()]
        if not keys and not ids:
            return jsonify({'issues': []})

        auth_string = f"{JIRA_EMAIL}:{JIRA_TOKEN}"
        auth_bytes = auth_string.encode('ascii')
        auth_base64 = base64.b64encode(auth_bytes).decode('ascii')

        headers = {
            'Authorization': f'Basic {auth_base64}',
            'Accept': 'application/json',
            'Content-Type': 'application/json'
        }

        team_field_id = resolve_team_field_id(headers)
        epic_link_field_id = resolve_epic_link_field_id(headers)

        fields_list = [
            'summary',
            'status',
            'issuetype',
            'customfield_10004',
            'parent'
        ]
        if epic_link_field_id and epic_link_field_id not in fields_list:
            fields_list.append(epic_link_field_id)
        if team_field_id and team_field_id not in fields_list:
            fields_list.append(team_field_id)
        if JIRA_TEAM_FALLBACK_FIELD_ID not in fields_list:
            fields_list.append(JIRA_TEAM_FALLBACK_FIELD_ID)

        issues = []
        if keys:
            unique_keys = sorted({str(k).strip() for k in keys if str(k).strip()})
            issues.extend(fetch_issues_by_keys(unique_keys, headers, fields_list))

        if ids:
            unique_ids = sorted({str(i).strip() for i in ids if str(i).strip()})
            jql = f'id in ({",".join(unique_ids)})'
            payload = {
                'jql': jql,
                'startAt': 0,
                'maxResults': len(unique_ids),
                'fields': fields_list
            }
            response = jira_search_request(headers, payload)
            if response.status_code == 200:
                data = response.json() or {}
                issues.extend(data.get('issues', []) or [])
            else:
                print(f'❌ Lookup fetch error: {response.status_code} {response.text}')

        snapshots = []
        for issue in issues:
            snapshot = build_issue_snapshot(issue, team_field_id, epic_link_field_id)
            snapshot['id'] = issue.get('id')
            snapshots.append(snapshot)

        return jsonify({'issues': snapshots})
    except Exception as e:
        print(f'❌ Issue lookup error: {e}')
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Failed to lookup issues', 'message': str(e)}), 500


@app.route('/api/scenario', methods=['GET', 'POST'])
def scenario_planner():
    """Scenario planner endpoint."""
    if request.method == 'GET':
        if not SCENARIO_CACHE.get('data'):
            return jsonify({'error': 'No scenario cached'}), 404
        return jsonify(SCENARIO_CACHE)

    try:
        payload = request.get_json(silent=True) or {}
        config_payload = payload.get('config') or {}
        filters = payload.get('filters') or {}

        sprint_label = resolve_sprint_label(filters.get('sprint'))
        quarter_start, quarter_end = quarter_dates_from_label(sprint_label)
        start_date = parse_iso_date(config_payload.get('start_date')) or quarter_start or date.today()
        quarter_end_date = parse_iso_date(config_payload.get('quarter_end_date')) or quarter_end or (start_date + timedelta(days=90))

        scenario_config = ScenarioConfig(
            start_date=start_date,
            quarter_end_date=quarter_end_date,
            sp_to_weeks=2.0,
            team_sizes={},
            vacation_weeks={},
            sickleave_buffer=0.0,
            wip_limit=1,
            lane_mode=config_payload.get('lane_mode', 'team'),
        )

        auth_string = f"{JIRA_EMAIL}:{JIRA_TOKEN}"
        auth_bytes = auth_string.encode('ascii')
        auth_base64 = base64.b64encode(auth_bytes).decode('ascii')
        headers = {
            'Authorization': f'Basic {auth_base64}',
            'Accept': 'application/json',
            'Content-Type': 'application/json'
        }

        team_field_id = resolve_team_field_id(headers)
        epic_link_field_id = resolve_epic_link_field_id(headers)

        fields_list = [
            'summary',
            'status',
            'priority',
            'issuetype',
            'assignee',
            'updated',
            'customfield_10004',
            'parent'
        ]
        if epic_link_field_id and epic_link_field_id not in fields_list:
            fields_list.append(epic_link_field_id)
        if team_field_id and team_field_id not in fields_list:
            fields_list.append(team_field_id)
        if JIRA_TEAM_FALLBACK_FIELD_ID not in fields_list:
            fields_list.append(JIRA_TEAM_FALLBACK_FIELD_ID)

        search_query = (filters.get('search') or '').strip().lower()
        team_filter_ids = {t for t in (filters.get('teams') or []) if t}
        scenario_jql = build_scenario_jql(filters)
        issues_raw = fetch_issues_by_jql(scenario_jql, headers, fields_list)

        issues = []
        issue_keys = []
        issue_by_key = {}
        team_names = set()
        epic_keys = set()
        for issue in issues_raw:
            fields = issue.get('fields', {}) or {}
            raw_team = None
            if fields.get(JIRA_TEAM_FALLBACK_FIELD_ID) is not None:
                raw_team = fields.get(JIRA_TEAM_FALLBACK_FIELD_ID)
            elif team_field_id and fields.get(team_field_id) is not None:
                raw_team = fields.get(team_field_id)
            team_name = extract_team_name(raw_team) if raw_team is not None else None
            team_id = None
            if raw_team is not None:
                team_value = build_team_value(raw_team)
                if isinstance(team_value, dict):
                    team_id = team_value.get('id')

            epic_key = None
            if epic_link_field_id and fields.get(epic_link_field_id):
                epic_key = fields.get(epic_link_field_id)
            elif fields.get('parent') and fields['parent'].get('key') and \
                    fields['parent'].get('fields', {}).get('issuetype', {}).get('name', '').lower() == 'epic':
                epic_key = fields['parent'].get('key')

            assignee = fields.get('assignee') or {}
            issue_type = (fields.get('issuetype') or {}).get('name') or ''
            story_points = fields.get('customfield_10004')
            priority = (fields.get('priority') or {}).get('name')
            status = (fields.get('status') or {}).get('name')
            issue_obj = Issue(
                key=issue.get('key'),
                summary=fields.get('summary') or '',
                issue_type=issue_type,
                team=team_name,
                assignee=assignee.get('displayName'),
                story_points=story_points,
                priority=priority,
                status=status,
                epic_key=epic_key,
                team_id=team_id,
            )
            if issue_obj.key:
                issues.append(issue_obj)
                issue_keys.append(issue_obj.key)
                if team_name:
                    team_names.add(team_name)
                if epic_key:
                    epic_keys.add(epic_key)
                issue_by_key[issue_obj.key] = {
                    'key': issue_obj.key,
                    'summary': issue_obj.summary,
                    'type': issue_type,
                    'team': team_name,
                    'team_id': team_id,
                    'assignee': issue_obj.assignee,
                    'sp': story_points,
                    'priority': priority,
                    'status': status,
                    'epicKey': epic_key,
                }

        dependencies = collect_dependencies(issue_keys, headers)
        dependency_edges = {}
        edge_list = []
        edge_set = set()
        dependency_snapshots = {}
        for deps in dependencies.values():
            for dep in deps:
                if dep.get('key'):
                    dependency_snapshots[dep['key']] = dep

        for issue_key, deps in dependencies.items():
            for dep in deps:
                prereq_key = dep.get('prereqKey')
                dependent_key = dep.get('dependentKey')
                if not prereq_key or not dependent_key:
                    continue
                category = dep.get('category')
                edge_type = 'dependency' if category == 'dependency' else 'block' if category == 'block' else None
                if edge_type is None:
                    continue
                if prereq_key == dependent_key:
                    continue
                edge_id = (prereq_key, dependent_key, edge_type)
                if edge_id in edge_set:
                    continue
                edge_set.add(edge_id)
                edge_list.append({'from': prereq_key, 'to': dependent_key, 'type': edge_type})
                if edge_type in ('dependency', 'block'):
                    dependency_edges.setdefault(dependent_key, []).append(prereq_key)

        def matches_search(entry):
            if not search_query:
                return True
            key = (entry.get('key') or '').lower()
            summary = (entry.get('summary') or '').lower()
            return search_query in key or search_query in summary

        focus_keys = [
            key for key, entry in issue_by_key.items()
            if matches_search(entry) and (not team_filter_ids or entry.get('team_id') in team_filter_ids)
        ]

        adjacency = {}
        for edge in edge_list:
            adjacency.setdefault(edge['from'], set()).add(edge['to'])
            adjacency.setdefault(edge['to'], set()).add(edge['from'])

        if len(team_filter_ids) == 1:
            focus_keys = [key for key in focus_keys if adjacency.get(key)]

        focus_set = set(focus_keys)
        context_keys = set()
        for key in focus_set:
            context_keys.update(adjacency.get(key, set()))
        context_keys -= focus_set

        included_keys = focus_set | context_keys
        if not focus_set and not search_query:
            included_keys = set(issue_by_key.keys())
            focus_set = set(included_keys)

        for key in context_keys:
            if key in issue_by_key:
                continue
            snapshot = dependency_snapshots.get(key)
            if not snapshot:
                continue
            issue_by_key[key] = {
                'key': snapshot.get('key'),
                'summary': snapshot.get('summary') or '',
                'type': snapshot.get('issuetype'),
                'team': snapshot.get('teamName'),
                'team_id': snapshot.get('teamId'),
                'assignee': None,
                'sp': snapshot.get('storyPoints'),
                'priority': snapshot.get('priority'),
                'status': snapshot.get('status'),
                'epicKey': snapshot.get('epicKey'),
            }
            if snapshot.get('teamName'):
                team_names.add(snapshot.get('teamName'))
            if snapshot.get('epicKey'):
                epic_keys.add(snapshot.get('epicKey'))

        capacity_details = {}
        if sprint_label:
            capacity_keys = {}
            for name in team_names:
                normalized = normalize_capacity_team_name(name)
                if normalized:
                    capacity_keys[name] = normalized
            capacity_sizes, capacity_details = fetch_capacity_team_sizes(
                sprint_label,
                headers,
                team_names=sorted(set(capacity_keys.values()))
            )
            scenario_config.team_sizes = {
                name: capacity_sizes.get(norm)
                for name, norm in capacity_keys.items()
                if capacity_sizes.get(norm) is not None
            }

        epic_summary_by_key = {}
        if epic_keys:
            epic_issues = fetch_issues_by_keys(sorted(epic_keys), headers, ['summary'])
            for epic in epic_issues:
                fields = epic.get('fields') or {}
                epic_summary_by_key[epic.get('key')] = fields.get('summary')

        jira_base_url = (JIRA_URL or '').rstrip('/')

        capacity_by_team = {}
        if sprint_label:
            for team_name in sorted(team_names):
                normalized = normalize_capacity_team_name(team_name)
                detail = capacity_details.get(normalized) if normalized else None
                size = detail.get('watchers') if detail else None
                capacity_by_team[team_name] = {
                    'size': size,
                    'capacityIssueKey': detail.get('issue_key') if detail else None,
                    'watchersCount': detail.get('watchers') if detail else None
                }

        issue_objs = []
        for key in included_keys:
            entry = issue_by_key.get(key)
            if not entry:
                continue
            issue_objs.append(Issue(
                key=entry.get('key'),
                summary=entry.get('summary') or '',
                issue_type=entry.get('type') or '',
                team=entry.get('team'),
                assignee=entry.get('assignee'),
                story_points=entry.get('sp'),
                priority=entry.get('priority'),
                status=entry.get('status'),
                epic_key=entry.get('epicKey'),
                team_id=entry.get('team_id'),
            ))

        scheduled_list, scheduled_map = schedule_issues(issue_objs, dependency_edges, scenario_config)
        scheduled_by_key = {item.key: item for item in scheduled_list}
        slack, critical = compute_slack(scheduled_map, dependency_edges, scenario_config.quarter_end_date)
        if app.debug:
            blocked_edges = [edge for edge in edge_list if edge.get('type') == 'block']
            for edge in blocked_edges[:20]:
                prereq_key = edge.get('from')
                dependent_key = edge.get('to')
                assert prereq_key != dependent_key
                prereq_item = scheduled_by_key.get(prereq_key)
                dependent_item = scheduled_by_key.get(dependent_key)
                prereq_start = prereq_item.start_date.isoformat() if prereq_item and prereq_item.start_date else None
                prereq_end = prereq_item.end_date.isoformat() if prereq_item and prereq_item.end_date else None
                dependent_start = dependent_item.start_date.isoformat() if dependent_item and dependent_item.start_date else None
                dependent_end = dependent_item.end_date.isoformat() if dependent_item and dependent_item.end_date else None
                print(
                    "scenario blocked_by edge",
                    {
                        "prereqKey": prereq_key,
                        "dependentKey": dependent_key,
                        "prereqStart": prereq_start,
                        "prereqEnd": prereq_end,
                        "dependentStart": dependent_start,
                        "dependentEnd": dependent_end,
                    },
                )

        total_weeks = max(1.0, (scenario_config.quarter_end_date - scenario_config.start_date).days / 7.0)
        lane_usage = {}
        for item in scheduled_list:
            if item.duration_weeks is None:
                continue
            lane_usage[item.lane] = lane_usage.get(item.lane, 0.0) + item.duration_weeks

        bottleneck_lanes = sorted(lane_usage.keys(), key=lambda lane: lane_usage[lane], reverse=True)[:3]
        late_items = []
        unschedulable = []

        for item in scheduled_list:
            if item.key in slack:
                item.slack_weeks = slack[item.key]
                item.is_critical = item.key in critical
            if item.end_date and item.end_date > scenario_config.quarter_end_date:
                item.is_late = True
                late_items.append(item.key)
            if item.scheduled_reason != 'scheduled' and item.scheduled_reason != 'already_done':
                unschedulable.append(item.key)

        response_issues = []
        for key in sorted(included_keys):
            entry = issue_by_key.get(key)
            item = scheduled_by_key.get(key)
            if not entry:
                continue
            epic_key = entry.get('epicKey')
            response_issues.append({
                'key': key,
                'summary': entry.get('summary'),
                'type': entry.get('type'),
                'team': entry.get('team'),
                'team_id': entry.get('team_id'),
                'assignee': entry.get('assignee'),
                'sp': entry.get('sp'),
                'priority': entry.get('priority'),
                'status': entry.get('status'),
                'epicKey': epic_key,
                'epicSummary': epic_summary_by_key.get(epic_key),
                'start': item.start_date.isoformat() if item and item.start_date else None,
                'end': item.end_date.isoformat() if item and item.end_date else None,
                'blockedBy': item.blocked_by if item else [],
                'scheduledReason': item.scheduled_reason if item else 'context_only',
                'durationWeeks': item.duration_weeks if item else None,
                'slackWeeks': item.slack_weeks if item else None,
                'isCritical': item.is_critical if item else False,
                'isLate': item.is_late if item else False,
                'isContext': key in context_keys,
                'url': f'{jira_base_url}/browse/{key}' if jira_base_url else None,
            })

        result = {
            'generatedAt': datetime.now().isoformat(),
            'jira_base_url': jira_base_url,
            'config': {
                'start_date': scenario_config.start_date.isoformat(),
                'quarter_end_date': scenario_config.quarter_end_date.isoformat(),
                'sp_to_weeks': scenario_config.sp_to_weeks,
                'wip_limit': scenario_config.wip_limit,
                'sickleave_buffer': scenario_config.sickleave_buffer,
                'lane_mode': scenario_config.lane_mode,
                'sprint': sprint_label
            },
            'summary': {
                'critical_path': critical,
                'bottleneck_lanes': bottleneck_lanes,
                'late_items': late_items,
                'unschedulable': unschedulable,
                'deadline_met': len(late_items) == 0 and len(unschedulable) == 0,
            },
            'issues': response_issues,
            'dependencies': [edge for edge in edge_list if edge['from'] in included_keys and edge['to'] in included_keys],
            'capacity_by_team': capacity_by_team,
            'focus_set': {
                'focused_issue_keys': sorted(focus_set),
                'context_issue_keys': sorted(context_keys),
            },
        }

        SCENARIO_CACHE['generatedAt'] = result['generatedAt']
        SCENARIO_CACHE['data'] = result

        return jsonify(result)
    except Exception as e:
        print(f'❌ Scenario error: {e}')
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Failed to compute scenario', 'message': str(e)}), 500


def fetch_stats_for_sprint(sprint_name, headers, team_field_id, team_ids=None):
    """Fetch stories for a sprint and aggregate delivery stats by team/project."""
    base_jql = STATS_JQL_BASE or f'project in ("{JIRA_PRODUCT_PROJECT}","{JIRA_TECH_PROJECT}")'
    base_jql = strip_sprint_clause(base_jql)
    stats_team_ids = team_ids or get_stats_team_ids()
    if stats_team_ids and not re.search(r'"Team\[Team\]"\s+in\s*\(', base_jql, flags=re.IGNORECASE) and \
            not re.search(r'"Team\[Team\]"\s*=\s*', base_jql, flags=re.IGNORECASE):
        if len(stats_team_ids) == 1:
            base_jql = add_clause_to_jql(base_jql, f'"Team[Team]" = "{stats_team_ids[0]}"')
        else:
            quoted = ', '.join(f'"{team_id}"' for team_id in stats_team_ids)
            base_jql = add_clause_to_jql(base_jql, f'"Team[Team]" in ({quoted})')
    jql = add_clause_to_jql(base_jql, f'Sprint in ("{sprint_name}")')
    if STATS_JQL_ORDER_BY and not re.search(r'order\s+by', jql, flags=re.IGNORECASE):
        jql = f"{jql} {STATS_JQL_ORDER_BY}"

    fields_list = [
        'status',
        'project',
        'priority',
        'customfield_10004'  # Story Points
    ]
    if team_field_id and team_field_id not in fields_list:
        fields_list.append(team_field_id)
    if JIRA_TEAM_FALLBACK_FIELD_ID not in fields_list:
        fields_list.append(JIRA_TEAM_FALLBACK_FIELD_ID)

    page_size = 250
    start_at = 0
    collected_issues = []
    total_issues = None

    while True:
        payload = {
            'jql': jql,
            'startAt': start_at,
            'maxResults': page_size,
            'fields': fields_list
        }

        response = jira_search_request(headers, payload)
        if response.status_code != 200:
            return None, response

        data = response.json()
        total_issues = data.get('total', total_issues)
        issues = data.get('issues', [])
        if not issues:
            break

        collected_issues.extend(issues)
        if len(issues) < payload['maxResults']:
            break
        start_at += len(issues)
        if total_issues is not None and start_at >= total_issues:
            break

    def normalize_status(value):
        return (value or '').strip().lower()

    def parse_points(value):
        try:
            if value is None:
                return 0.0
            return float(value)
        except Exception:
            return 0.0

    teams = {}
    projects_summary = {}
    totals = {
        'done': 0,
        'incomplete': 0,
        'killed': 0,
        'donePoints': 0.0,
        'incompletePoints': 0.0
    }

    for issue in collected_issues:
        fields = issue.get('fields', {}) or {}
        status_name = (fields.get('status') or {}).get('name', '')
        status_value = normalize_status(status_name)
        is_done = status_value == 'done'
        is_killed = status_value == 'killed'
        priority_name = (fields.get('priority') or {}).get('name', '') or 'Unspecified'

        points = parse_points(fields.get('customfield_10004'))
        project_name = (fields.get('project') or {}).get('name')
        project_key = (fields.get('project') or {}).get('key')
        project_label = project_name or project_key or 'Unknown Project'
        project_bucket = classify_project(project_label)

        raw_team = None
        if fields.get(JIRA_TEAM_FALLBACK_FIELD_ID) is not None:
            raw_team = fields.get(JIRA_TEAM_FALLBACK_FIELD_ID)
        elif team_field_id and fields.get(team_field_id) is not None:
            raw_team = fields.get(team_field_id)
        team_ids = extract_team_ids(raw_team)
        if stats_team_ids:
            if not team_ids:
                continue
            if not any(team_id in stats_team_ids for team_id in team_ids):
                continue

        team_payload = build_team_value(raw_team) if raw_team is not None else {}
        team_id = None
        team_name = None
        if isinstance(team_payload, dict):
            team_id = team_payload.get('id') or (team_ids[0] if team_ids else None)
            team_name = team_payload.get('name')
        if not team_name:
            team_name = extract_team_name(raw_team)

        team_key = team_id or team_name or 'unknown'
        if team_key not in teams:
            teams[team_key] = {
                'id': team_id,
                'name': team_name or 'Unknown Team',
                'done': 0,
                'incomplete': 0,
                'killed': 0,
                'donePoints': 0.0,
                'incompletePoints': 0.0,
                'projects': {},
                'priorities': {}
            }

        team_entry = teams[team_key]
        if priority_name not in team_entry['priorities']:
            team_entry['priorities'][priority_name] = {'done': 0, 'incomplete': 0, 'killed': 0}
        if project_bucket not in team_entry['projects']:
            team_entry['projects'][project_bucket] = {
                'done': 0,
                'incomplete': 0,
                'killed': 0,
                'donePoints': 0.0,
                'incompletePoints': 0.0,
                'priorities': {}
            }

        if project_bucket not in projects_summary:
            projects_summary[project_bucket] = {
                'done': 0,
                'incomplete': 0,
                'killed': 0,
                'donePoints': 0.0,
                'incompletePoints': 0.0
            }

        if is_killed:
            team_entry['killed'] += 1
            team_entry['priorities'][priority_name]['killed'] += 1
            if priority_name not in team_entry['projects'][project_bucket]['priorities']:
                team_entry['projects'][project_bucket]['priorities'][priority_name] = {'done': 0, 'incomplete': 0, 'killed': 0}
            team_entry['projects'][project_bucket]['priorities'][priority_name]['killed'] += 1
            team_entry['projects'][project_bucket]['killed'] += 1
            projects_summary[project_bucket]['killed'] += 1
            totals['killed'] += 1
            continue

        if is_done:
            team_entry['done'] += 1
            team_entry['donePoints'] += points
            team_entry['priorities'][priority_name]['done'] += 1
            if priority_name not in team_entry['projects'][project_bucket]['priorities']:
                team_entry['projects'][project_bucket]['priorities'][priority_name] = {'done': 0, 'incomplete': 0, 'killed': 0}
            team_entry['projects'][project_bucket]['priorities'][priority_name]['done'] += 1
            team_entry['projects'][project_bucket]['done'] += 1
            team_entry['projects'][project_bucket]['donePoints'] += points
            projects_summary[project_bucket]['done'] += 1
            projects_summary[project_bucket]['donePoints'] += points
            totals['done'] += 1
            totals['donePoints'] += points
        else:
            team_entry['incomplete'] += 1
            team_entry['incompletePoints'] += points
            team_entry['priorities'][priority_name]['incomplete'] += 1
            if priority_name not in team_entry['projects'][project_bucket]['priorities']:
                team_entry['projects'][project_bucket]['priorities'][priority_name] = {'done': 0, 'incomplete': 0, 'killed': 0}
            team_entry['projects'][project_bucket]['priorities'][priority_name]['incomplete'] += 1
            team_entry['projects'][project_bucket]['incomplete'] += 1
            team_entry['projects'][project_bucket]['incompletePoints'] += points
            projects_summary[project_bucket]['incomplete'] += 1
            projects_summary[project_bucket]['incompletePoints'] += points
            totals['incomplete'] += 1
            totals['incompletePoints'] += points

    sorted_teams = sorted(
        teams.values(),
        key=lambda t: (t['name'] or '').lower()
    )

    stats_payload = {
        'sprint': sprint_name,
        'totals': totals,
        'projects': projects_summary,
        'teams': sorted_teams
    }
    return stats_payload, None


def build_missing_info_scope_clause(team_ids, component_name):
    clauses = []
    component_name = (component_name or '').strip()
    team_ids = [t.strip() for t in (team_ids or []) if t and str(t).strip()]

    if component_name:
        clauses.append(f'component = "{component_name}"')
    if team_ids:
        if len(team_ids) == 1:
            clauses.append(f'"Team[Team]" = "{team_ids[0]}"')
        else:
            quoted = ', '.join(f'"{t}"' for t in team_ids)
            clauses.append(f'"Team[Team]" in ({quoted})')

    if not clauses:
        return ''
    if len(clauses) == 1:
        return clauses[0]
    return f"({ ' OR '.join(clauses) })"


@app.route('/api/missing-info', methods=['GET'])
def get_missing_info():
    """Find stories under epics in a given sprint that are missing key planning fields (sprint/SP/team)."""
    try:
        sprint = request.args.get('sprint', '').strip()
        if not sprint:
            return jsonify({'error': 'Missing required query param: sprint'}), 400

        # Prepare authorization
        auth_string = f"{JIRA_EMAIL}:{JIRA_TOKEN}"
        auth_bytes = auth_string.encode('ascii')
        auth_base64 = base64.b64encode(auth_bytes).decode('ascii')

        headers = {
            'Authorization': f'Basic {auth_base64}',
            'Accept': 'application/json',
            'Content-Type': 'application/json'
        }

        # Resolve fields
        team_field_id = resolve_team_field_id(headers)
        epic_link_field_id = resolve_epic_link_field_id(headers)

        scope_clause = build_missing_info_scope_clause(MISSING_INFO_TEAM_IDS, MISSING_INFO_COMPONENT)

        # 1) Fetch epics that are in the sprint (future sprint planning), scoped by component/team.
        epic_jql = f'Sprint = {sprint} AND issuetype = Epic'
        if scope_clause:
            epic_jql = add_clause_to_jql(epic_jql, scope_clause)
        epic_jql = add_clause_to_jql(epic_jql, 'status not in ("Killed","Done","Incomplete")')
        epic_jql = add_clause_to_jql(epic_jql, f'project in ("{JIRA_PRODUCT_PROJECT}","{JIRA_TECH_PROJECT}")')

        epic_fields = ['summary', 'status', 'assignee', 'parent']
        if team_field_id:
            epic_fields.append(team_field_id)
        if JIRA_TEAM_FALLBACK_FIELD_ID not in epic_fields:
            epic_fields.append(JIRA_TEAM_FALLBACK_FIELD_ID)

        epics_resp = jira_search_request(headers, {
            'jql': epic_jql,
            'startAt': 0,
            'maxResults': 250,
            'fields': epic_fields
        })
        if epics_resp.status_code != 200:
            return jsonify({'error': 'Failed to fetch epics for missing-info scan', 'details': epics_resp.text}), 502

        epics_data = epics_resp.json() or {}
        epic_issues = epics_data.get('issues', []) or []
        epic_keys = [e.get('key') for e in epic_issues if e.get('key')]
        if not epic_keys:
            return jsonify({'issues': [], 'epics': [], 'count': 0})

        # 2) Fetch stories under those epics, regardless of story sprint (to catch missing Sprint field).
        story_fields = [
            'summary',
            'status',
            'priority',
            'issuetype',
            'assignee',
            'updated',
            'customfield_10004',  # Story Points
            'customfield_10101',  # Sprint
            'parent'
        ]
        if epic_link_field_id and epic_link_field_id not in story_fields:
            story_fields.append(epic_link_field_id)
        if team_field_id and team_field_id not in story_fields:
            story_fields.append(team_field_id)
        if JIRA_TEAM_FALLBACK_FIELD_ID not in story_fields:
            story_fields.append(JIRA_TEAM_FALLBACK_FIELD_ID)

        missing = []
        batch_size = 40
        for start in range(0, len(epic_keys), batch_size):
            batch = epic_keys[start:start + batch_size]

            link_clause = f'"Epic Link" in ({",".join(batch)})'
            parent_clause = f'parent in ({",".join(batch)})'
            # Important: do NOT scope stories by component/team here because the whole point is to
            # find stories missing those fields. We only scope epics, then pull every story under them.
            story_jql = f'({link_clause} OR {parent_clause}) AND issuetype = Story AND status not in (Killed, Done, Postponed)'

            start_at = 0
            while True:
                resp = jira_search_request(headers, {
                    'jql': story_jql,
                    'startAt': start_at,
                    'maxResults': 250,
                    'fields': story_fields
                })
                if resp.status_code != 200:
                    break

                data = resp.json() or {}
                issues = data.get('issues', []) or []
                if not issues:
                    break

                for issue in issues:
                    fields = issue.get('fields', {}) or {}
                    status = (fields.get('status') or {}).get('name') or ''
                    if str(status).strip().lower() == 'postponed':
                        continue

                    # team enrichment
                    raw_team = None
                    if fields.get(JIRA_TEAM_FALLBACK_FIELD_ID) is not None:
                        raw_team = fields.get(JIRA_TEAM_FALLBACK_FIELD_ID)
                    elif team_field_id and fields.get(team_field_id) is not None:
                        raw_team = fields.get(team_field_id)
                    if raw_team is not None:
                        team_name = extract_team_name(raw_team)
                        fields['team'] = build_team_value(raw_team)
                        fields['teamName'] = team_name
                        fields['teamId'] = fields['team'].get('id') if isinstance(fields['team'], dict) else None

                    # epic link
                    epic_key = None
                    if epic_link_field_id and fields.get(epic_link_field_id):
                        epic_key = fields.get(epic_link_field_id)
                    elif fields.get('parent') and fields['parent'].get('key') and \
                            fields['parent'].get('fields', {}).get('issuetype', {}).get('name', '').lower() == 'epic':
                        epic_key = fields['parent'].get('key')
                    if epic_key:
                        fields['epicKey'] = epic_key

                    sp = fields.get('customfield_10004')
                    try:
                        sp_num = float(sp) if sp not in (None, '', []) else 0.0
                    except Exception:
                        sp_num = 0.0
                    has_sp = sp_num > 0

                    sprint_value = fields.get('customfield_10101')
                    has_sprint = bool(sprint_value)
                    has_team = bool(fields.get('teamName'))

                    missing_fields = []
                    if not has_sprint:
                        missing_fields.append('Sprint')
                    if not has_sp:
                        missing_fields.append('Story Points')
                    if not has_team:
                        missing_fields.append('Team')

                    if not missing_fields:
                        continue

                    assignee = fields.get('assignee') or {}
                    priority = fields.get('priority') or {}
                    issuetype = fields.get('issuetype') or {}
                    missing.append({
                        'id': issue.get('id'),
                        'key': issue.get('key'),
                        'fields': {
                            'summary': fields.get('summary'),
                            'status': {'name': status} if status else None,
                            'priority': {'name': priority.get('name')} if priority else None,
                            'issuetype': {'name': issuetype.get('name')} if issuetype else None,
                            'assignee': {'displayName': assignee.get('displayName')} if assignee else None,
                            'updated': fields.get('updated'),
                            'customfield_10004': fields.get('customfield_10004'),
                            'customfield_10101': fields.get('customfield_10101'),
                            'team': fields.get('team'),
                            'teamName': fields.get('teamName'),
                            'teamId': fields.get('teamId'),
                            'epicKey': fields.get('epicKey'),
                            'missingFields': missing_fields
                        }
                    })

                start_at += len(issues)
                total = data.get('total')
                if total is not None and start_at >= total:
                    break
                if len(issues) < 250:
                    break

        response = jsonify({'issues': missing, 'count': len(missing), 'epicCount': len(epic_keys)})
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        return response
    except Exception as e:
        print(f'❌ Missing-info error: {e}')
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Failed to compute missing-info', 'message': str(e)}), 500


@app.route('/api/tasks', methods=['GET'])
def get_tasks():
    """Fetch tasks from Jira API."""
    return fetch_tasks(include_team_name=False)


@app.route('/api/tasks-with-team-name', methods=['GET'])
def get_tasks_with_team_name():
    """Fetch tasks with team name derived from Jira Team field."""
    return fetch_tasks(include_team_name=True)


@app.route('/api/stats', methods=['GET'])
def get_completed_sprint_stats():
    """Fetch cached delivery stats for a completed sprint."""
    sprint_name = request.args.get('sprint', '').strip()
    team_id = request.args.get('team', '').strip()
    team_ids_raw = request.args.get('teamIds', '').strip()
    refresh = request.args.get('refresh', '').lower() == 'true'

    if not sprint_name:
        return jsonify({'error': 'Missing sprint name'}), 400

    base_jql = STATS_JQL_BASE or JQL_QUERY
    team_ids = []
    if team_ids_raw:
        team_ids = [t.strip() for t in team_ids_raw.split(',') if t.strip()]
    elif team_id:
        team_ids = [team_id]
    else:
        team_ids = get_stats_team_ids()
    cache_key = build_stats_cache_key(sprint_name, base_jql, team_ids)
    cache_data = load_stats_cache()
    if not refresh and cache_key in cache_data:
        cached_payload = cache_data.get(cache_key, {})
        response = {
            'cached': True,
            'generatedAt': cached_payload.get('generatedAt'),
            'data': cached_payload.get('data')
        }
        return jsonify(response)

    auth_string = f"{JIRA_EMAIL}:{JIRA_TOKEN}"
    auth_bytes = auth_string.encode('ascii')
    auth_base64 = base64.b64encode(auth_bytes).decode('ascii')

    headers = {
        'Authorization': f'Basic {auth_base64}',
        'Accept': 'application/json',
        'Content-Type': 'application/json'
    }

    team_field_id = resolve_team_field_id(headers)
    stats_payload, error_response = fetch_stats_for_sprint(sprint_name, headers, team_field_id, team_ids=team_ids or None)
    if error_response is not None:
        return jsonify({
            'error': 'Failed to fetch stats',
            'details': error_response.text
        }), error_response.status_code

    generated_at = datetime.now().isoformat()
    cache_data[cache_key] = {
        'generatedAt': generated_at,
        'data': stats_payload
    }
    save_stats_cache(cache_data)

    return jsonify({
        'cached': False,
        'generatedAt': generated_at,
        'data': stats_payload
    })


@app.route('/api/stats-example', methods=['GET'])
def get_stats_example():
    """Serve example stats payload if available."""
    if not STATS_EXAMPLE_FILE or not os.path.exists(STATS_EXAMPLE_FILE):
        return jsonify({'error': 'Stats example file not found'}), 404
    try:
        with open(STATS_EXAMPLE_FILE, 'r') as f:
            payload = json.load(f)
        return jsonify(payload)
    except Exception as e:
        return jsonify({'error': 'Failed to load stats example', 'details': str(e)}), 500


@app.route('/api/boards', methods=['GET'])
def get_boards():
    """Fetch available boards from Jira API"""
    try:
        auth_string = f"{JIRA_EMAIL}:{JIRA_TOKEN}"
        auth_bytes = auth_string.encode('ascii')
        auth_base64 = base64.b64encode(auth_bytes).decode('ascii')

        headers = {
            'Authorization': f'Basic {auth_base64}',
            'Accept': 'application/json',
            'Content-Type': 'application/json'
        }

        print(f'\n📋 Fetching all boards...')

        # Get boards from Jira Agile API
        response = requests.get(
            f'{JIRA_URL}/rest/agile/1.0/board',
            headers=headers,
            params={'maxResults': 100},
            timeout=30
        )

        print(f'Boards Response Status: {response.status_code}')

        if response.status_code != 200:
            error_text = response.text
            print(f'❌ Error Response: {error_text}')
            return jsonify({
                'error': f'Jira API error: {response.status_code}',
                'details': error_text
            }), response.status_code

        data = response.json()
        boards = data.get('values', [])

        # Format boards
        formatted_boards = []
        for board in boards:
            formatted_boards.append({
                'id': board.get('id'),
                'name': board.get('name'),
                'type': board.get('type'),
                'location': board.get('location', {})
            })

        print(f'✅ Found {len(formatted_boards)} boards')

        success_response = jsonify({'boards': formatted_boards})
        success_response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        success_response.headers['Pragma'] = 'no-cache'
        success_response.headers['Expires'] = '0'
        return success_response

    except Exception as e:
        print(f'❌ Exception: {str(e)}')
        import traceback
        traceback.print_exc()
        error_response = jsonify({
            'error': 'Failed to fetch boards from Jira',
            'message': str(e)
        })
        error_response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        return error_response, 500


@app.route('/api/sprints', methods=['GET'])
def get_sprints():
    """Fetch available sprints - uses cache if valid, otherwise fetches from Jira"""
    try:
        force_refresh = request.args.get('refresh', '').lower() == 'true'

        formatted_sprints = []

        # Check if we should use cache
        if not force_refresh and is_cache_valid():
            cache_data = load_sprints_cache()
            if cache_data and 'sprints' in cache_data:
                formatted_sprints = cache_data['sprints']
                print(f'📦 Loaded {len(formatted_sprints)} sprints from cache')

        # If no valid cache or force refresh, fetch from Jira
        if not formatted_sprints or force_refresh:
            if force_refresh:
                print('🔄 Force refresh requested')

            formatted_sprints = fetch_sprints_from_jira()

            # Save to cache
            if formatted_sprints:
                save_sprints_cache(formatted_sprints)

        print(f'✅ Total quarterly sprints: {len(formatted_sprints)}')

        success_response = jsonify({'sprints': formatted_sprints})
        success_response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        success_response.headers['Pragma'] = 'no-cache'
        success_response.headers['Expires'] = '0'
        return success_response

    except Exception as e:
        print(f'❌ Exception: {str(e)}')
        import traceback
        traceback.print_exc()
        error_response = jsonify({
            'error': 'Failed to fetch sprints from Jira',
            'message': str(e)
        })
        error_response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        return error_response, 500


@app.route('/api/config', methods=['GET'])
def get_config():
    """Get public configuration"""
    return jsonify({
        'jiraUrl': JIRA_URL,
        'capacityProject': CAPACITY_PROJECT
    })


@app.route('/api/capacity', methods=['GET'])
def get_capacity():
    """Get estimated team capacity for a sprint."""
    sprint_name = request.args.get('sprint', '').strip()
    debug = request.args.get('debug', '').lower() in ('1', 'true', 'yes')
    team_param = request.args.get('teams', '').strip()
    team_names = [s.strip() for s in team_param.split(',') if s.strip()]
    if not sprint_name:
        return jsonify({'error': 'Sprint name is required'}), 400

    if not CAPACITY_PROJECT:
        return jsonify({
            'enabled': False,
            'capacities': {}
        })

    try:
        auth_string = f"{JIRA_EMAIL}:{JIRA_TOKEN}"
        auth_bytes = auth_string.encode('ascii')
        auth_base64 = base64.b64encode(auth_bytes).decode('ascii')

        headers = {
            'Authorization': f'Basic {auth_base64}',
            'Accept': 'application/json',
            'Content-Type': 'application/json'
        }

        payload, error_message = fetch_capacity_for_sprint(sprint_name, headers, debug=debug, team_names=team_names)
        if error_message:
            return jsonify({'error': error_message}), 500
        return jsonify(payload)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/planned-capacity', methods=['GET'])
def get_planned_capacity():
    """Alias endpoint for planned capacity."""
    return get_capacity()


@app.route('/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({
        'status': 'OK',
        'message': 'Jira proxy server is running'
    })


@app.route('/api/test', methods=['GET'])
def test_connection():
    """Test Jira connection with simple query"""
    try:
        auth_string = f"{JIRA_EMAIL}:{JIRA_TOKEN}"
        auth_bytes = auth_string.encode('ascii')
        auth_base64 = base64.b64encode(auth_bytes).decode('ascii')

        headers = {
            'Authorization': f'Basic {auth_base64}',
            'Accept': 'application/json',
            'Content-Type': 'application/json'
        }

        # Simple test query - just get any 5 issues from PRODUCT project
        test_payload = {
            'jql': 'project = PRODUCT ORDER BY created DESC',
            'maxResults': 5,
            'fields': ['summary', 'status', 'priority']
        }

        print(f'\n🧪 Testing Jira connection...')
        print(f'Test JQL: {test_payload["jql"]}')

        response = jira_search_request(headers, test_payload)

        print(f'Test Response Status: {response.status_code}')

        if response.status_code != 200:
            return jsonify({
                'status': 'error',
                'code': response.status_code,
                'message': response.text
            }), response.status_code

        data = response.json()
        return jsonify({
            'status': 'success',
            'message': f'Connection OK! Found {len(data.get("issues", []))} test issues',
            'sample_issue': data.get('issues', [{}])[0].get('key', 'N/A') if data.get('issues') else None
        })

    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500


@app.route('/api/debug-fields', methods=['GET'])
def debug_fields():
    """Debug endpoint to see all fields of a single task"""
    try:
        auth_string = f"{JIRA_EMAIL}:{JIRA_TOKEN}"
        auth_bytes = auth_string.encode('ascii')
        auth_base64 = base64.b64encode(auth_bytes).decode('ascii')

        headers = {
            'Authorization': f'Basic {auth_base64}',
            'Accept': 'application/json',
            'Content-Type': 'application/json'
        }

        # Get one issue with ALL fields
        payload = {
            'jql': JQL_QUERY,
            'maxResults': 1,
            'fields': ['*all']
        }

        print(f'\n🔍 Fetching all fields for debugging...')

        response = jira_search_request(headers, payload)

        if response.status_code != 200:
            return jsonify({
                'error': f'Jira API error: {response.status_code}',
                'details': response.text
            }), response.status_code

        data = response.json()

        if data.get('issues') and len(data['issues']) > 0:
            issue = data['issues'][0]
            fields = issue.get('fields', {})

            # Look for Story Points in customfields
            customfields = {}
            for key, value in fields.items():
                if key.startswith('customfield_') and value is not None:
                    customfields[key] = value

            return jsonify({
                'issue_key': issue.get('key'),
                'all_customfields': customfields,
                'fields_keys': list(fields.keys())
            })
        else:
            return jsonify({
                'error': 'No issues found',
                'jql': JQL_QUERY
            }), 404

    except Exception as e:
        return jsonify({
            'error': str(e)
        }), 500


@app.route('/api/tasks-fields', methods=['GET'])
def get_tasks_fields():
    """Return all available fields and values for issues matching JQL_QUERY."""
    try:
        auth_string = f"{JIRA_EMAIL}:{JIRA_TOKEN}"
        auth_bytes = auth_string.encode('ascii')
        auth_base64 = base64.b64encode(auth_bytes).decode('ascii')

        headers = {
            'Authorization': f'Basic {auth_base64}',
            'Accept': 'application/json',
            'Content-Type': 'application/json'
        }

        limit = request.args.get('limit', '5')
        try:
            limit_value = max(1, min(int(limit), 50))
        except ValueError:
            limit_value = 5

        payload = {
            'jql': JQL_QUERY,
            'maxResults': limit_value,
            'fields': ['*all']
        }

        print(f'\n🔍 Fetching all fields for {limit_value} issues...')

        response = jira_search_request(headers, payload)

        if response.status_code != 200:
            return jsonify({
                'error': f'Jira API error: {response.status_code}',
                'details': response.text
            }), response.status_code

        data = response.json()
        issues = data.get('issues', [])

        return jsonify({
            'total': data.get('total'),
            'returned': len(issues),
            'issues': issues
        })

    except Exception as e:
        return jsonify({
            'error': str(e)
        }), 500


@app.route('/api/export-excel', methods=['POST'])
def export_excel():
    """Export selected tasks to Excel file"""
    try:
        data = request.get_json()
        tasks = data.get('tasks', [])

        if not tasks:
            return jsonify({'error': 'No tasks provided'}), 400

        print(f'\n📊 Exporting {len(tasks)} tasks to Excel...')

        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Sprint Tasks'

        # Define header style
        header_fill = PatternFill(start_color='107C41', end_color='107C41', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_alignment = Alignment(horizontal='center', vertical='center')

        # Add headers
        headers = ['ID', 'Subject', 'Story Points']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Add data
        for row_num, task in enumerate(tasks, 2):
            ws.cell(row=row_num, column=1, value=task.get('key', ''))
            ws.cell(row=row_num, column=2, value=task.get('summary', ''))
            ws.cell(row=row_num, column=3, value=task.get('storyPoints', 0))

        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 15

        # Align Story Points column to center
        for row in range(2, len(tasks) + 2):
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        print(f'✅ Excel file generated successfully')

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'sprint_tasks_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
        )

    except Exception as e:
        print(f'❌ Export error: {str(e)}')
        import traceback
        traceback.print_exc()
        return jsonify({
            'error': 'Failed to export to Excel',
            'message': str(e)
        }), 500


if __name__ == '__main__':
    args = parse_args()

    # Apply CLI overrides while keeping env defaults as fallbacks
    if args.jira_url:
        JIRA_URL = args.jira_url
    if args.jira_email:
        JIRA_EMAIL = args.jira_email
    if args.jira_token:
        JIRA_TOKEN = args.jira_token
    if args.jira_query:
        JQL_QUERY = args.jira_query
    if args.server_port:
        SERVER_PORT = args.server_port

    # Validate configuration
    if not JIRA_URL or not JIRA_EMAIL or not JIRA_TOKEN:
        print('\n❌ ERROR: JIRA_URL, JIRA_EMAIL and JIRA_TOKEN must be set via environment or CLI!')
        print('📝 Please copy .env.example to .env, fill in your credentials, or pass them as flags.\n')
        exit(1)

    print('\n🚀 Jira Proxy Server starting...')
    print(f'📧 Using email: {JIRA_EMAIL}')
    print(f'🔗 Jira URL: {JIRA_URL}')
    print(f'📊 Board ID: {JIRA_BOARD_ID}')
    print(f'📝 JQL Query: {JQL_QUERY[:80]}...' if len(JQL_QUERY) > 80 else f'📝 JQL Query: {JQL_QUERY}')
    print(f'💾 Cache expires after: {CACHE_EXPIRY_HOURS} hours')
    print('\n📋 Endpoints:')
    print(f'   • http://localhost:{SERVER_PORT}/api/tasks - Get sprint tasks')
    print(f'   • http://localhost:{SERVER_PORT}/api/tasks-with-team-name - Get sprint tasks with team names')
    print(f'   • http://localhost:{SERVER_PORT}/api/dependencies - Get issue dependencies (POST)')
    print(f'   • http://localhost:{SERVER_PORT}/api/issues/lookup?keys=KEY-1,KEY-2 - Lookup issues (GET)')
    print(f'   • http://localhost:{SERVER_PORT}/api/scenario - Scenario planner (GET/POST)')
    print(f'   • http://localhost:{SERVER_PORT}/api/stats?sprint=2025Q3 - Get completed sprint stats (cached)')
    print(f'   • http://localhost:{SERVER_PORT}/api/stats-example - Get example completed sprint stats')
    print(f'   • http://localhost:{SERVER_PORT}/api/sprints - Get available sprints (cached)')
    print(f'   • http://localhost:{SERVER_PORT}/api/sprints?refresh=true - Force refresh sprints cache')
    print(f'   • http://localhost:{SERVER_PORT}/api/planned-capacity?sprint=2025Q3 - Get planned team capacity')
    print(f'   • http://localhost:{SERVER_PORT}/api/boards - Get all boards (to find board ID)')
    print(f'   • http://localhost:{SERVER_PORT}/api/test - Test connection')
    print(f'   • http://localhost:{SERVER_PORT}/health - Health check')
    print('\n✅ Server ready! Open jira-dashboard.html in your browser\n')

    app.run(host='0.0.0.0', port=SERVER_PORT, debug=True)
